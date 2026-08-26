"""OfficeBench evaluation functions with upstream-style output organization."""

import datetime
import difflib
import glob
import json
import os
import re
from email import policy
from email.parser import BytesParser
from typing import Any, Dict, List

import icalendar
import openpyxl
import pytz
from docx import Document
from PyPDF2 import PdfReader


PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))


def _resolve_file_path(testbed_dir: str, file_path: str) -> str:
    if os.path.isabs(file_path):
        return file_path
    testbed_path = os.path.join(testbed_dir, file_path)
    if os.path.exists(testbed_path):
        return testbed_path
    project_path = os.path.join(PROJECT_ROOT, file_path)
    if file_path.replace("\\", "/").startswith(("benchmark/", "dataset/", "scripts/")) and os.path.exists(project_path):
        return project_path
    return testbed_path


def _is_number(string: str) -> bool:
    try:
        float(string)
        return True
    except ValueError:
        return False


def _read_excel(file_path: str) -> str:
    sheet = openpyxl.load_workbook(file_path).active
    content_string = ""
    for row in sheet.iter_rows():
        for cell in row:
            value = cell.value if cell.value is not None else "[Empty Cell]"
            content_string += f"({cell.row}, {cell.column}): {value}\t"
        content_string += "\n"
    return content_string


def _read_word(file_path: str) -> str:
    doc = Document(file_path)
    return "\n".join([paragraph.text for paragraph in doc.paragraphs])


def _read_pdf(file_path: str) -> str:
    reader = PdfReader(file_path)
    text = ""
    for page in reader.pages:
        text += page.extract_text() or ""
    return text


def _get_email_content(msg) -> str:
    if msg.is_multipart():
        parts = []
        for part in msg.iter_parts():
            if part.get_content_type() in {"text/plain", "text/html"}:
                charset = part.get_content_charset() or "utf-8"
                parts.append(part.get_payload(decode=True).decode(charset, errors="replace"))
        return "\n".join(parts)
    charset = msg.get_content_charset() or "utf-8"
    return msg.get_payload(decode=True).decode(charset, errors="replace")


def _list_emails(username: str, testbed_dir: str, first_n_letters: int = -1) -> str:
    email_folder = os.path.join(testbed_dir, "emails", username)
    os.makedirs(email_folder, exist_ok=True)
    email_files = glob.glob(os.path.join(email_folder, "*.eml"))
    message = ""
    for email_file in email_files:
        with open(email_file, "rb") as f:
            email_obj = BytesParser(policy=policy.default).parsebytes(f.read())
        email_name = os.path.basename(email_file)
        body = _get_email_content(email_obj)
        message += f"Email ID: {email_name}\n"
        message += f"From: {email_obj['From']}\n"
        message += f"To: {email_obj['To']}\n"
        message += f"Subject: {email_obj['Subject']}\n"
        if first_n_letters != -1:
            message += f"Content: {body[:first_n_letters] + '...'}\n"
        else:
            message += f"Content: {body}\n"
        message += "-" * 50 + "\n"
    return message


def _evaluate_contain_text(content: str, args: Dict[str, Any]) -> bool:
    content = content.lower()
    for keyword in args["keywords"]:
        keyword = str(keyword).lower()
        if _is_number(keyword):
            content = content.replace(",", "")
        if keyword not in content:
            return False
    return True


def evaluate_contain(testbed_dir: str, args: Dict[str, Any]) -> bool:
    doc_type = args["doc_type"]
    if doc_type == "email":
        username = args["username"]
        email_contents = ""
        direct = os.path.join(testbed_dir, "emails", username)
        lower = os.path.join(testbed_dir, "emails", username.lower())
        if os.path.exists(direct):
            email_contents = _list_emails(username, testbed_dir, -1)
        elif os.path.exists(lower):
            email_contents = _list_emails(username.lower(), testbed_dir, -1)
        else:
            email_accounts = glob.glob(os.path.join(testbed_dir, "emails", "*"))
            for email_account in email_accounts:
                if username.lower() in os.path.basename(email_account).lower():
                    email_contents = _list_emails(os.path.basename(email_account), testbed_dir, -1)
                    break
        return _evaluate_contain_text(email_contents, args)

    file_path = _resolve_file_path(testbed_dir, args["file"])
    if not os.path.exists(file_path) or not os.path.isfile(file_path):
        return False
    if doc_type == "xlsx":
        content = _read_excel(file_path)
    elif doc_type in {"txt", "ics"}:
        with open(file_path, "r", encoding="utf-8") as f:
            content = f.read()
    elif doc_type in {"doc", "docx"}:
        content = _read_word(file_path)
    elif doc_type == "pdf":
        content = _read_pdf(file_path)
    else:
        raise ValueError(f"Not implemented doc type: {doc_type}")
    return _evaluate_contain_text(content, args)


def evaluate_not_contain(testbed_dir: str, args: Dict[str, Any]) -> bool:
    return not evaluate_contain(testbed_dir, args)


def evaluate_file_exist(testbed_dir: str, args: Dict[str, Any]) -> bool:
    return os.path.exists(os.path.join(testbed_dir, args["file"]))


def evaluate_file_not_exist(testbed_dir: str, args: Dict[str, Any]) -> bool:
    return not os.path.exists(os.path.join(testbed_dir, args["file"]))


def _helper_diff_contain_text(input_content: str, output_content: str, args: Dict[str, Any]) -> bool:
    if input_content == output_content:
        return False
    diff = difflib.unified_diff(input_content.split("\n"), output_content.split("\n"), n=0)
    diff_text = "\n".join(list(diff))
    for matches in args["keywords"]:
        if str(matches) not in diff_text:
            return False
    return True


def evaluate_diff_contain_text(testbed_dir: str, args: Dict[str, Any]) -> bool:
    doc_type = args["doc_type"]
    input_file = _resolve_file_path(testbed_dir, args["input_file"])
    output_file = _resolve_file_path(testbed_dir, args["output_file"])
    if doc_type == "xlsx":
        input_content = _read_excel(input_file)
        output_content = _read_excel(output_file)
    elif doc_type == "doc":
        input_content = _read_word(input_file)
        output_content = _read_word(output_file)
    else:
        raise ValueError(f"Not implemented doc type: {doc_type}")
    return _helper_diff_contain_text(input_content, output_content, args)


def evaluate_excel_cell_value(testbed_dir: str, args: Dict[str, Any]) -> bool:
    file_path = _resolve_file_path(testbed_dir, args["file"])
    if not os.path.exists(file_path):
        return False
    content = _read_excel(file_path)
    for match in args["matches"]:
        pattern = f"({match['row']}, {match['col']}): {match['value']}"
        if pattern not in content:
            return False
    return True


def evaluate_excel_cell_comparator(testbed_dir: str, args: Dict[str, Any]) -> bool:
    file_path = _resolve_file_path(testbed_dir, args["file"])
    content = _read_excel(file_path)
    for match in args["matches"]:
        pattern = r"\({}, {}\): (\w+)\t".format(match["row"], match["col"])
        found = re.search(pattern, content)
        if not found:
            return False
        value = found.group(1)
        comparator = eval(match["comparator"])  # noqa: S307
        if not comparator(value):
            return False
    return True


def evaluate_calendar_no_overlap(testbed_dir: str, args: Dict[str, Any]) -> bool:
    username = args["username"]
    calendar_file = os.path.join(testbed_dir, "calendar", f"{username}.ics")
    calendar = icalendar.Calendar.from_ical(open(calendar_file, "rb").read())
    utc = pytz.UTC

    def is_naive(dt: datetime.datetime) -> bool:
        return dt.tzinfo is None

    def proc_dt(dt: datetime.datetime) -> datetime.datetime:
        return utc.localize(dt) if is_naive(dt) else dt

    calendar.subcomponents.sort(key=lambda x: proc_dt(x.get("dtstart").dt))
    events = []
    for component in calendar.walk():
        if component.name == "VEVENT":
            events.append(component)
    for i in range(len(events) - 1):
        if proc_dt(events[i].get("dtend").dt) > proc_dt(events[i + 1].get("dtstart").dt):
            return False
    return True


def evaluate_exact_match(testbed_dir: str, args: Dict[str, Any]) -> bool:
    result_path = _resolve_file_path(testbed_dir, args["result_file"])
    expected_path = _resolve_file_path(testbed_dir, args["expected_file"])
    if not os.path.exists(result_path):
        return False
    doc_type = args["doc_type"]
    if doc_type != "xlsx":
        if doc_type in {"txt", "ics"}:
            with open(result_path, "r", encoding="utf-8") as f:
                result_content = f.read()
            with open(expected_path, "r", encoding="utf-8") as f:
                expected_content = f.read()
        elif doc_type == "doc":
            result_content = _read_word(result_path)
            expected_content = _read_word(expected_path)
        elif doc_type == "pdf":
            result_content = _read_pdf(result_path)
            expected_content = _read_pdf(expected_path)
        else:
            raise ValueError(f"Not implemented doc type: {doc_type}")
        return result_content == expected_content

    result_sheet = openpyxl.load_workbook(result_path).active
    expected_sheet = openpyxl.load_workbook(expected_path).active
    for row in result_sheet.iter_rows():
        for cell in row:
            if cell.value != expected_sheet.cell(row=cell.row, column=cell.column).value:
                return False
    for row in expected_sheet.iter_rows():
        for cell in row:
            if cell.value != result_sheet.cell(row=cell.row, column=cell.column).value:
                return False
    return True


EVALUATION_MAP = {
    "evaluate_contain": evaluate_contain,
    "evaluate_not_contain": evaluate_not_contain,
    "evaluate_file_exist": evaluate_file_exist,
    "evaluate_file_not_exist": evaluate_file_not_exist,
    "evaluate_diff_contain_text": evaluate_diff_contain_text,
    "evaluate_excel_cell_value": evaluate_excel_cell_value,
    "evaluate_excel_cell_comparator": evaluate_excel_cell_comparator,
    "evaluate_exact_match": evaluate_exact_match,
    "evaluate_calendar_no_overlap": evaluate_calendar_no_overlap,
}


def _run_evaluation_items(testbed_dir: str, evaluation_items: List[Dict[str, Any]]) -> bool:
    for item in evaluation_items:
        fn_name = item.get("function", "")
        args = item.get("args", {})
        fn = EVALUATION_MAP.get(fn_name)
        if fn is None:
            return False
        try:
            if not bool(fn(testbed_dir, args)):
                return False
        except Exception:
            return False
    return True


def evaluate_officebench_task(testbed_dir: str, evaluation_items: List[Dict[str, Any]]) -> Dict[str, Any]:
    is_pass = bool(evaluation_items) and _run_evaluation_items(testbed_dir, evaluation_items)
    return {
        "score": 1.0 if is_pass else 0.0,
        "is_pass": is_pass,
        "actual_score": int(is_pass),
        "max_score": 1,
        "percentage": 1.0 if is_pass else 0.0,
        "criteria_results": [],
    }


def load_officebench_subtask_config(dataset_root: str, task_id: str, subtask_id: str) -> Dict[str, Any]:
    config_path = os.path.join(dataset_root, "tasks", task_id, "subtasks", f"{subtask_id}.json")
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


def evaluate_officebench_output(dataset_root: str, task_id: str, subtask_id: str, output_dir: str) -> bool:
    config = load_officebench_subtask_config(dataset_root, task_id, subtask_id)
    eval_config = config.get("evaluation", [])
    return _run_evaluation_items(output_dir, eval_config)


def discover_officebench_tasks(dataset_root: str) -> List[tuple[str, str]]:
    all_tasks_info: List[tuple[str, str]] = []
    pattern = os.path.join(dataset_root, "tasks", "*", "subtasks", "*.json")
    for config_filepath in glob.glob(pattern):
        parts = config_filepath.split(os.sep)
        task_id = parts[-3]
        subtask_id = os.path.splitext(parts[-1])[0]
        all_tasks_info.append((task_id, subtask_id))
    return sorted(all_tasks_info, key=lambda x: tuple(map(int, x[0].split("-"))) + (int(x[1]),))
