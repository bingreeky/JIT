"""OdysseyBench evaluation helpers."""

from __future__ import annotations

import difflib
import glob
import os
from email import policy
from email.parser import BytesParser
from typing import Any, Dict, List

import icalendar
import openpyxl
import pytz
from docx import Document
from PyPDF2 import PdfReader


PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", ".."))


def _eval_is_number(string: str) -> bool:
    try:
        float(string)
        return True
    except (TypeError, ValueError):
        return False


def _iter_candidate_paths(testbed_dir: str, relative_path: str) -> List[str]:
    rel = str(relative_path or "").strip()
    if not rel:
        return []

    rel = rel.replace("\\", "/")
    if rel == "/testbed":
        return [os.path.normpath(testbed_dir)]
    if rel.startswith("/testbed/"):
        rel = rel[len("/testbed/"):]
    if os.path.isabs(rel):
        return [os.path.normpath(rel)]

    rel = rel.removeprefix("./").removeprefix("/")
    candidates = [os.path.join(testbed_dir, rel)]
    if rel.startswith(("benchmark/", "dataset/", "scripts/")):
        candidates.append(os.path.join(PROJECT_ROOT, rel))
    if not rel.startswith(("data/", "emails/", "calendar/")):
        candidates.extend(
            [
                os.path.join(testbed_dir, "data", rel),
                os.path.join(testbed_dir, "emails", rel),
                os.path.join(testbed_dir, "calendar", rel),
            ]
        )

    unique: List[str] = []
    seen = set()
    for candidate in candidates:
        norm = os.path.normpath(candidate)
        if norm not in seen:
            seen.add(norm)
            unique.append(norm)
    return unique


def _resolve_eval_path(testbed_dir: str, relative_path: str) -> str:
    candidates = _iter_candidate_paths(testbed_dir, relative_path)
    for candidate in candidates:
        if os.path.exists(candidate):
            return candidate
    return candidates[0] if candidates else os.path.join(testbed_dir, str(relative_path or ""))


def _infer_doc_type(file_path: str) -> str:
    return os.path.splitext(file_path)[1].lower().lstrip(".")


def _read_excel(file_path: str, sheet_name: str | None = None) -> str:
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    content = ""
    for row in worksheet.iter_rows():
        for cell in row:
            value = cell.value if cell.value is not None else "[Empty Cell]"
            content += f"({cell.row}, {cell.column}): {value}\t"
        content += "\n"
    return content


def _read_word(file_path: str) -> str:
    doc = Document(file_path)
    return "\n".join(paragraph.text for paragraph in doc.paragraphs)


def _read_pdf(file_path: str) -> str:
    reader = PdfReader(file_path)
    return "".join(page.extract_text() or "" for page in reader.pages)


def _read_text(file_path: str) -> str:
    with open(file_path, "r", encoding="utf-8") as handle:
        return handle.read()


def _get_email_content(msg) -> str:
    if msg.is_multipart():
        parts = []
        for part in msg.iter_parts():
            if part.get_content_type() in {"text/plain", "text/html"}:
                charset = part.get_content_charset() or "utf-8"
                payload = part.get_payload(decode=True) or b""
                parts.append(payload.decode(charset, errors="replace"))
        return "\n".join(parts)
    charset = msg.get_content_charset() or "utf-8"
    payload = msg.get_payload(decode=True) or b""
    return payload.decode(charset, errors="replace")


def _list_emails(username: str, testbed_dir: str, first_n_letters: int = -1) -> str:
    email_folder = os.path.join(testbed_dir, "emails", username)
    os.makedirs(email_folder, exist_ok=True)
    email_files = glob.glob(os.path.join(email_folder, "*.eml"))
    message = ""
    for email_file in email_files:
        with open(email_file, "rb") as handle:
            email_obj = BytesParser(policy=policy.default).parsebytes(handle.read())
        email_name = os.path.basename(email_file)
        body = _get_email_content(email_obj)
        message += f"Email ID: {email_name}\n"
        message += f"From: {email_obj['From']}\n"
        message += f"To: {email_obj['To']}\n"
        message += f"Subject: {email_obj['Subject']}\n"
        message += f"Content: {body if first_n_letters == -1 else body[:first_n_letters] + '...'}\n"
        message += "-" * 50 + "\n"
    return message


def _read_document(file_path: str, doc_type: str, sheet_name: str | None = None) -> str:
    if doc_type == "xlsx":
        return _read_excel(file_path, sheet_name=sheet_name)
    if doc_type in {"txt", "ics"}:
        return _read_text(file_path)
    if doc_type in {"doc", "docx"}:
        return _read_word(file_path)
    if doc_type == "pdf":
        return _read_pdf(file_path)
    raise ValueError(f"Unsupported doc type: {doc_type}")


def _evaluate_contain_text(content: str, keywords: List[Any]) -> bool:
    haystack = content.lower()
    normalized_haystack = haystack.replace(",", "")
    for keyword in keywords:
        needle = str(keyword).lower()
        lookup = normalized_haystack if _eval_is_number(needle) else haystack
        if needle not in lookup:
            return False
    return True


def evaluate_contain(testbed_dir: str, args: Dict[str, Any]) -> bool:
    doc_type = str(args["doc_type"]).lower()
    if doc_type == "email":
        username = args["username"]
        email_contents = ""
        direct = os.path.join(testbed_dir, "emails", username)
        lower = os.path.join(testbed_dir, "emails", username.lower())
        if os.path.exists(direct):
            email_contents = _list_emails(username, testbed_dir, -1)
        elif os.path.exists(lower):
            email_contents = _list_emails(username.lower(), testbed_dir, -1)
        else:
            for email_account in glob.glob(os.path.join(testbed_dir, "emails", "*")):
                if username.lower() in os.path.basename(email_account).lower():
                    email_contents = _list_emails(os.path.basename(email_account), testbed_dir, -1)
                    break
        return _evaluate_contain_text(email_contents, args["keywords"])

    file_path = _resolve_eval_path(testbed_dir, args["file"])
    if not os.path.isfile(file_path):
        return False
    content = _read_document(file_path, doc_type, sheet_name=args.get("sheet"))
    return _evaluate_contain_text(content, args["keywords"])


def evaluate_not_contain(testbed_dir: str, args: Dict[str, Any]) -> bool:
    return not evaluate_contain(testbed_dir, args)


def evaluate_file_exist(testbed_dir: str, args: Dict[str, Any]) -> bool:
    return os.path.exists(_resolve_eval_path(testbed_dir, args["file"]))


def evaluate_file_not_exist(testbed_dir: str, args: Dict[str, Any]) -> bool:
    return not evaluate_file_exist(testbed_dir, args)


def evaluate_diff_contain_text(testbed_dir: str, args: Dict[str, Any]) -> bool:
    doc_type = str(args["doc_type"]).lower()
    input_file = _resolve_eval_path(testbed_dir, args["input_file"])
    output_file = _resolve_eval_path(testbed_dir, args["output_file"])
    if not os.path.exists(input_file) or not os.path.exists(output_file):
        return False
    input_content = _read_document(input_file, doc_type, sheet_name=args.get("sheet"))
    output_content = _read_document(output_file, doc_type, sheet_name=args.get("sheet"))
    if input_content == output_content:
        return False
    diff = difflib.unified_diff(input_content.split("\n"), output_content.split("\n"), n=0)
    diff_text = "\n".join(list(diff))
    return all(str(keyword) in diff_text for keyword in args["keywords"])


def evaluate_excel_cell_value(testbed_dir: str, args: Dict[str, Any]) -> bool:
    file_path = _resolve_eval_path(testbed_dir, args["file"])
    if not os.path.exists(file_path):
        return False
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook[args["sheet"]] if args.get("sheet") else workbook.active
    for match in args["matches"]:
        value = worksheet.cell(row=int(match["row"]), column=int(match["col"])).value
        if str(value) != str(match["value"]):
            return False
    return True


def evaluate_excel_cell_comparator(testbed_dir: str, args: Dict[str, Any]) -> bool:
    file_path = _resolve_eval_path(testbed_dir, args["file"])
    if not os.path.exists(file_path):
        return False
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook[args["sheet"]] if args.get("sheet") else workbook.active
    for match in args["matches"]:
        value = worksheet.cell(row=int(match["row"]), column=int(match["col"])).value
        comparator = eval(match["comparator"])  # noqa: S307
        if not comparator("" if value is None else str(value)):
            return False
    return True


def evaluate_excel_sheet_exist(testbed_dir: str, args: Dict[str, Any]) -> bool:
    file_path = _resolve_eval_path(testbed_dir, args["file"])
    if not os.path.exists(file_path):
        return False
    workbook = openpyxl.load_workbook(file_path)
    return str(args["sheet"]) in workbook.sheetnames


def evaluate_calendar_no_overlap(testbed_dir: str, args: Dict[str, Any]) -> bool:
    username = args["username"]
    calendar_file = _resolve_eval_path(testbed_dir, os.path.join("calendar", f"{username}.ics"))
    if not os.path.exists(calendar_file):
        return False
    calendar = icalendar.Calendar.from_ical(open(calendar_file, "rb").read())
    utc = pytz.UTC

    def proc_dt(dt):
        if dt.tzinfo is None:
            return utc.localize(dt)
        return dt

    events = [component for component in calendar.walk() if component.name == "VEVENT"]
    events.sort(key=lambda component: proc_dt(component.get("dtstart").dt))
    for idx in range(len(events) - 1):
        if proc_dt(events[idx].get("dtend").dt) > proc_dt(events[idx + 1].get("dtstart").dt):
            return False
    return True


def evaluate_exact_match(testbed_dir: str, args: Dict[str, Any]) -> bool:
    result_path = _resolve_eval_path(testbed_dir, args["result_file"])
    expected_path = _resolve_eval_path(testbed_dir, args["expected_file"])
    if not os.path.exists(result_path) or not os.path.exists(expected_path):
        return False

    doc_type = str(args["doc_type"]).lower()
    if doc_type == "xlsx":
        result_book = openpyxl.load_workbook(result_path)
        expected_book = openpyxl.load_workbook(expected_path)
        result_sheet = result_book[args["sheet"]] if args.get("sheet") else result_book.active
        expected_sheet = expected_book[args["sheet"]] if args.get("sheet") else expected_book.active

        max_row = max(result_sheet.max_row, expected_sheet.max_row)
        max_col = max(result_sheet.max_column, expected_sheet.max_column)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                if result_sheet.cell(row=row, column=col).value != expected_sheet.cell(row=row, column=col).value:
                    return False
        return True

    result_content = _read_document(result_path, doc_type, sheet_name=args.get("sheet"))
    expected_content = _read_document(expected_path, doc_type, sheet_name=args.get("sheet"))
    return result_content == expected_content


def evaluate_keyword_exist(testbed_dir: str, args: Dict[str, Any]) -> bool:
    file_path = _resolve_eval_path(testbed_dir, args["file"])
    if not os.path.exists(file_path):
        return False
    doc_type = str(args.get("doc_type") or _infer_doc_type(file_path)).lower()
    content = _read_document(file_path, doc_type, sheet_name=args.get("sheet"))
    return _evaluate_contain_text(content, args["keywords"])


def evaluate_doc_contain(testbed_dir: str, args: Dict[str, Any]) -> bool:
    file_path = _resolve_eval_path(testbed_dir, args["file"])
    if not os.path.exists(file_path):
        return False
    doc_type = str(args.get("doc_type") or _infer_doc_type(file_path)).lower()
    content = _read_document(file_path, doc_type, sheet_name=args.get("sheet"))
    return _evaluate_contain_text(content, args["keywords"])


def evaluate_file_contains_keywords(testbed_dir: str, args: Dict[str, Any]) -> bool:
    file_path = _resolve_eval_path(testbed_dir, args["file"])
    if not os.path.exists(file_path):
        return False
    doc_type = str(args.get("doc_type") or _infer_doc_type(file_path)).lower()
    content = _read_document(file_path, doc_type, sheet_name=args.get("sheet")).lower()
    normalized = content.replace(",", "")

    keyword_groups = args.get("keywords", [])
    if keyword_groups and not isinstance(keyword_groups[0], list):
        keyword_groups = [keyword_groups]

    for group in keyword_groups:
        matched = True
        for keyword in group:
            needle = str(keyword).lower()
            lookup = normalized if _eval_is_number(needle) else content
            if needle not in lookup:
                matched = False
                break
        if matched:
            return True
    return False


EVALUATION_MAP = {
    "evaluate_calendar_no_overlap": evaluate_calendar_no_overlap,
    "evaluate_contain": evaluate_contain,
    "evaluate_diff_contain_text": evaluate_diff_contain_text,
    "evaluate_doc_contain": evaluate_doc_contain,
    "evaluate_exact_match": evaluate_exact_match,
    "evaluate_excel_cell_comparator": evaluate_excel_cell_comparator,
    "evaluate_excel_cell_value": evaluate_excel_cell_value,
    "evaluate_excel_sheet_exist": evaluate_excel_sheet_exist,
    "evaluate_file_contains_keywords": evaluate_file_contains_keywords,
    "evaluate_file_exist": evaluate_file_exist,
    "evaluate_file_not_exist": evaluate_file_not_exist,
    "evaluate_keyword_exist": evaluate_keyword_exist,
    "evaluate_not_contain": evaluate_not_contain,
}


def evaluate_odysseybench_task(testbed_dir: str, evaluation_items: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Evaluate OdysseyBench task artifacts inside a testbed workspace."""
    is_pass = bool(evaluation_items)
    criteria_results: List[Dict[str, Any]] = []
    for item in evaluation_items:
        fn_name = item.get("function", "")
        args = item.get("args", {})
        fn = EVALUATION_MAP.get(fn_name)
        if fn is None:
            criteria_results.append(
                {
                    "function": fn_name,
                    "args": args,
                    "passed": False,
                    "error": f"Unsupported evaluation function: {fn_name}",
                }
            )
            is_pass = False
            break
        try:
            passed = bool(fn(testbed_dir, args))
            criteria_results.append(
                {
                    "function": fn_name,
                    "args": args,
                    "passed": passed,
                }
            )
            if not passed:
                is_pass = False
                break
        except Exception as exc:
            criteria_results.append(
                {
                    "function": fn_name,
                    "args": args,
                    "passed": False,
                    "error": str(exc),
                }
            )
            is_pass = False
            break

    return {
        "score": 1.0 if is_pass else 0.0,
        "is_pass": is_pass,
        "actual_score": int(is_pass),
        "max_score": 1,
        "percentage": 1.0 if is_pass else 0.0,
        "criteria_results": criteria_results,
    }
